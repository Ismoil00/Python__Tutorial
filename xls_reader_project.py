import openpyxl as xl

wb = xl.load_workbook("transaction.xlsx")
sheet = wb["Лист1"]

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 1.5
    corrected_price_cell = sheet.cell(row, 5)
    corrected_price_cell.value = corrected_price

wb.save("transaction2.xlsx")